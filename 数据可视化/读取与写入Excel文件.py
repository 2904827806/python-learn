import pandas as pd
import numpy as np

"""读取Excel文件，pandas.read_excel(io，sheetname=0，header=0，names=None,
index_col=None，usecols=None)"""
#io 字符串，xls，xlsx文件路径或类文件对象
#sheetname，字符串列表或者整数列表，默认值为0，如果是字符串则用于工作表名称，
#整数为索引：sheet_name=【0，1，’sheet3'】表示第一个，第二个和名为，sheet3的
#sheet页中的数据作为DataFrame
#header=0 ,表示取第一行的名为列名，数据为除列名以外的数据
#index_col=None，指定列为索引列
"""
../:表示当前程序文件所在目录的上一级目录
./：表示当前程序文件所在的目录
/：表示当前程序文件的根目录
"""


#列1，读取Excel文件
pd.set_option('display.unicode.east_asian_width',True)#解决列不对齐
fill = r"C:\Users\29048\Desktop\工资.xls"
rs = pd.read_excel(fill,sheet_name=0,dtype=str) #读取Excel表中的数据
#print(rs)

#往excl中写入数据
def write_data():
    import xlrd2
    from openpyxl import Workbook
    wb = Workbook()
    sh0 = wb.create_sheet('数据', 0)
    sh2 = wb.create_sheet('墨涵')
    sh2.cell(1,2).value = '数据'
    del wb['Sheet']
    wr = xlrd2.open_workbook(fill)
    sh = wr.sheet_by_index(0)
    a = []
    for i in range(sh.nrows):
        b = []
        for j in range(sh.ncols):
            b.append(sh.cell_value(i, j))
        a.append(b)
    for c in a:
        sh0.append(c)
    wb.save('数据.xlsx')
ses = pd.read_excel('./数据.xlsx',sheet_name='墨涵') #根据指定Excel工作簿名称，读取数据
#print(ses)
#列2，通过行列索引读取指定行列数据
s2 = pd.read_excel(fill,header=0)
#index_col指定那一列为行索引
#header指定那一行为列索引

#读取指定列的参数 usecols = [n] 读取多列就可以往【】填写几个列名
s3 = pd.read_excel(fill,sheet_name=0,usecols=[1,3],header=None)
print(s3)

#将数据写入Excel文件中
s = {'a': [1, 2, 3, 4, 6], 'b': [2, 5, 8, 9, 6], 'c': [1, 3, 4, 6, 7], 'd': [7, 8, 9, 6, 5], 'f': [7, 4, 1, 2, 5]}
se = pd.DataFrame(s)
se.to_excel('data.xlsx',index=True,header=False) #往Excel写入数据


